import argparse
import json
import re
import subprocess
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts" / "regression"))

from course_selection_regression_lib import Account, RemoteClient, parse_csrf_token  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Course gender-limit regression suite")
    parser.add_argument("--base-url", default="http://127.0.0.1:8080")
    parser.add_argument("--admin-username", default="admin")
    parser.add_argument("--admin-password", default="1234qwer")
    parser.add_argument("--teacher-password", default="abc127!!!")
    parser.add_argument("--results-dir", default="scripts/regression/results")
    return parser


def refresh_accounts_from_export(client: RemoteClient, output_path: Path) -> dict[str, Account]:
    output_path.write_bytes(client.get_bytes("/admin/student-accounts/export"))
    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    accounts: dict[str, Account] = {}
    for row in rows[1:]:
        if not row or not row[1] or not row[4] or not row[6]:
            continue
        account = Account(
            username=str(row[4]).strip(),
            password=str(row[6]).strip(),
            student_no=str(row[1]).strip(),
            student_name=str(row[0]).strip(),
        )
        accounts[account.student_no] = account
    return accounts


def fetch_all_classes(client: RemoteClient, admin_token: str) -> list[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=admin_token)
    return payload.get("data", [])


def find_student_with_gender_and_account(
    client: RemoteClient,
    admin_token: str,
    accounts: dict[str, Account],
    expected_gender: str,
) -> dict:
    for class_row in fetch_all_classes(client, admin_token):
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            student_no = str(item.get("studentNo", "")).strip()
            if item.get("gender") == expected_gender and student_no in accounts:
                return {"student": item, "account": accounts[student_no]}
    raise RuntimeError(f"no {expected_gender} student with usable account found")


def fetch_student_by_id(client: RemoteClient, admin_token: str, student_id: int) -> dict:
    for class_row in fetch_all_classes(client, admin_token):
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            if int(item.get("id")) == student_id:
                return item
    raise RuntimeError(f"student not found after update: {student_id}")


def create_teacher(client: RemoteClient, csrf: str, prefix: str, password: str) -> dict:
    phone = f"19{int(time.time() * 1000) % 1_000_000_000:09d}"
    name = f"{prefix}-Teacher-{phone[-4:]}"
    client.post_form_quick(
        "/admin/teachers/add",
        [
            ("_csrf", csrf),
            ("name", name),
            ("password", password),
            ("phone", phone),
            ("accountType", "TEACHER"),
        ],
    )
    page, _ = client.get_text(f"/admin/teachers?phone={phone}")
    match = re.search(r"/admin/teachers/reset-password/(\d+)", page)
    if not match:
        raise RuntimeError("failed to create teacher")
    return {"id": int(match.group(1)), "username": phone, "password": password, "name": name}


def create_event(client: RemoteClient, csrf: str, event_name: str, round2_in_future: bool = True) -> int:
    now = datetime.now()
    if round2_in_future:
        round2_start = now + timedelta(minutes=21)
        round2_end = now + timedelta(hours=2)
    else:
        round2_start = now - timedelta(minutes=1)
        round2_end = now + timedelta(hours=2)
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now + timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", round2_start.strftime("%Y-%m-%dT%H:%M")),
        ("round2End", round2_end.strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form_quick("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"failed to create event: {final_url}")
    return int(match.group(1))


def save_event_students(client: RemoteClient, event_id: int, student_ids: list[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick(f"/admin/courses/{event_id}/students/save", fields)


def create_course(
    client: RemoteClient,
    event_id: int,
    csrf: str,
    name: str,
    *,
    gender_limit: str,
    teacher_id: int | None = None,
    capacity: int = 5,
) -> int:
    fields = [
        ("_csrf", csrf),
        ("name", name),
        ("description", f"{name} gender-limit regression"),
        ("capacityMode", "GLOBAL"),
        ("genderLimit", gender_limit),
        ("totalCapacity", str(capacity)),
    ]
    if teacher_id is not None:
        fields.append(("teacherId", str(teacher_id)))
    client.post_form_quick(f"/admin/courses/{event_id}/courses/save", fields)
    page, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    match = re.search(r'data-course-id="(\d+)"[^>]*data-course-name="' + re.escape(name) + '"', page)
    if not match:
        raise RuntimeError(f"failed to find created course id for {name}")
    return int(match.group(1))


def activate_course(client: RemoteClient, event_id: int, course_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/{event_id}/courses/{course_id}/activate", [("_csrf", csrf)])


def start_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])


def process_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/process", [("_csrf", csrf)])
    for _ in range(180):
        payload, _ = client.get_text(f"/admin/courses/events/{event_id}/lottery-status")
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            time.sleep(1)
            continue
        if data.get("status") and data.get("status") != "PROCESSING":
            return
        time.sleep(1)
    raise RuntimeError("round1 processing timeout")


def update_student_gender(client: RemoteClient, student: dict, csrf: str, new_gender: str) -> None:
    school_class = student.get("schoolClass") or {}
    fields = [
        ("_csrf", csrf),
        ("name", str(student.get("name", ""))),
        ("gender", new_gender),
        ("studentNo", str(student.get("studentNo", ""))),
        ("idCard", str(student.get("idCard", "") or "")),
        ("electiveClass", str(student.get("electiveClass", "") or "")),
        ("classId", str(school_class.get("id", ""))),
        ("studentStatus", str(student.get("studentStatus", "在籍"))),
    ]
    client.post_form_quick(f"/admin/students/edit/{student['id']}", fields)


def submit_round1_preference(account: Account, course_id: int, base_url: str) -> dict:
    client = RemoteClient(base_url)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    return client.api_request("POST", f"/api/student/courses/{course_id}/prefer?preference=1", token=token)


def confirm_round1_preference(account: Account, base_url: str) -> dict:
    client = RemoteClient(base_url)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    return client.api_request("POST", "/api/student/courses/confirm", token=token)


def fetch_student_my_selections(account: Account, base_url: str) -> list[dict]:
    client = RemoteClient(base_url)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/my-selections", token=token)
    return payload.get("data", [])


def fetch_student_courses(account: Account, base_url: str) -> list[dict]:
    client = RemoteClient(base_url)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/courses", token=token)
    return payload.get("data", [])


def trigger_finalize(account: Account, base_url: str) -> dict | None:
    client = RemoteClient(base_url)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/events/current", token=token)
    return payload.get("data")


def attempt_gender_limit_update(client: RemoteClient, event_id: int, course_id: int, course_name: str, csrf: str) -> tuple[str, str]:
    client.post_form_quick(
        f"/admin/courses/{event_id}/courses/save",
        [
            ("_csrf", csrf),
            ("courseId", str(course_id)),
            ("name", course_name),
            ("description", f"{course_name} guard test"),
            ("capacityMode", "GLOBAL"),
            ("genderLimit", "FEMALE_ONLY"),
            ("totalCapacity", "5"),
        ],
    )
    page, final_url = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    return page, final_url


def parse_course_gender_limit(detail_html: str, course_id: int) -> str | None:
    match = re.search(rf'data-course-id="{course_id}"[^>]*data-course-gender-limit="([^"]+)"', detail_html)
    return match.group(1) if match else None


def pick_selection(selections: list[dict], course_id: int) -> dict | None:
    for item in selections:
        if item.get("courseId") == course_id:
            return item
    return None


def get_confirmed_course_name(selections: list[dict]) -> str | None:
    for item in selections:
        if item.get("status") == "CONFIRMED":
            return item.get("courseName")
    return None


def summarize_courses(courses: list[dict]) -> dict[str, dict]:
    result: dict[str, dict] = {}
    for item in courses:
        result[str(item.get("name"))] = {
            "eligible": item.get("eligible"),
            "genderLimit": item.get("genderLimit"),
            "genderLimitLabel": item.get("genderLimitLabel"),
            "ineligibleMessage": item.get("ineligibleMessage"),
            "remaining": item.get("remaining"),
        }
    return result


def update_event_times_for_finalize(client: RemoteClient, csrf: str, event_id: int, event_name: str) -> None:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("id", str(event_id)),
        ("name", event_name),
        ("round1Start", (now - timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now - timedelta(hours=1, minutes=30)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now - timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M")),
    ]
    client.post_form_quick("/admin/courses/events/save", fields)


def run_edit_guard_scenario(
    admin_client: RemoteClient,
    admin_csrf: str,
    admin_token: str,
    accounts: dict[str, Account],
    base_url: str,
) -> dict:
    male_target = find_student_with_gender_and_account(admin_client, admin_token, accounts, "男")
    event_name = f"Gender-Edit-Guard-{datetime.now().strftime('%H%M%S')}"
    course_name = f"Guard-Course-{datetime.now().strftime('%H%M%S')}"
    event_id = create_event(admin_client, admin_csrf, event_name)
    save_event_students(admin_client, event_id, [int(male_target["student"]["id"])], admin_csrf)
    course_id = create_course(admin_client, event_id, admin_csrf, course_name, gender_limit="ALL")
    activate_course(admin_client, event_id, course_id, admin_csrf)
    start_round1(admin_client, event_id, admin_csrf)

    prefer_result = submit_round1_preference(male_target["account"], course_id, base_url)
    detail_html, detail_url = attempt_gender_limit_update(admin_client, event_id, course_id, course_name, admin_csrf)
    actual_gender_limit = parse_course_gender_limit(detail_html, course_id)

    passed = (
        prefer_result.get("code") == 200
        and detail_url.endswith("?tab=courses")
        and actual_gender_limit == "ALL"
    )
    return {
        "eventId": event_id,
        "courseId": course_id,
        "studentId": male_target["student"]["id"],
        "studentNo": male_target["account"].student_no,
        "preferResult": prefer_result,
        "detailUrl": detail_url,
        "actualGenderLimitAfterSaveAttempt": actual_gender_limit,
        "saveBlocked": actual_gender_limit == "ALL",
        "passed": passed,
    }


def run_pending_filter_scenario(
    admin_client: RemoteClient,
    admin_courses_csrf: str,
    admin_students_csrf: str,
    admin_token: str,
    accounts: dict[str, Account],
    base_url: str,
) -> dict:
    target = find_student_with_gender_and_account(admin_client, admin_token, accounts, "男")
    event_name = f"Gender-Pending-Filter-{datetime.now().strftime('%H%M%S')}"
    course_name = f"Pending-Filter-Course-{datetime.now().strftime('%H%M%S')}"
    event_id = create_event(admin_client, admin_courses_csrf, event_name)
    save_event_students(admin_client, event_id, [int(target["student"]["id"])], admin_courses_csrf)
    course_id = create_course(admin_client, event_id, admin_courses_csrf, course_name, gender_limit="MALE_ONLY")
    activate_course(admin_client, event_id, course_id, admin_courses_csrf)
    start_round1(admin_client, event_id, admin_courses_csrf)

    prefer_result = submit_round1_preference(target["account"], course_id, base_url)
    confirm_result = confirm_round1_preference(target["account"], base_url)
    update_student_gender(admin_client, target["student"], admin_students_csrf, "女")
    updated_student = fetch_student_by_id(admin_client, admin_token, int(target["student"]["id"]))
    process_round1(admin_client, event_id, admin_courses_csrf)
    selections_after = fetch_student_my_selections(target["account"], base_url)
    target_selection = pick_selection(selections_after, course_id)

    passed = (
        prefer_result.get("code") == 200
        and confirm_result.get("code") == 200
        and updated_student.get("gender") == "女"
        and target_selection is not None
        and target_selection.get("status") == "CANCELLED"
        and all(item.get("status") != "CONFIRMED" for item in selections_after)
    )
    return {
        "eventId": event_id,
        "courseId": course_id,
        "studentId": target["student"]["id"],
        "studentNo": target["account"].student_no,
        "preferResult": prefer_result,
        "confirmResult": confirm_result,
        "updatedStudentGender": updated_student.get("gender"),
        "targetSelectionAfterLottery": target_selection,
        "allSelectionsAfterLottery": selections_after,
        "passed": passed,
    }


def run_historical_invalid_scenario(
    admin_client: RemoteClient,
    admin_courses_csrf: str,
    admin_students_csrf: str,
    admin_token: str,
    teacher: dict,
    accounts: dict[str, Account],
    base_url: str,
) -> dict:
    target = find_student_with_gender_and_account(admin_client, admin_token, accounts, "男")

    def run_single(scenario_name: str, fallback: bool) -> dict:
        event_id = create_event(admin_client, admin_courses_csrf, scenario_name)
        save_event_students(admin_client, event_id, [int(target["student"]["id"])], admin_courses_csrf)
        invalid_course_name = f"{scenario_name}-invalid-course"
        invalid_course_id = create_course(
            admin_client,
            event_id,
            admin_courses_csrf,
            invalid_course_name,
            gender_limit="MALE_ONLY",
            teacher_id=int(teacher["id"]),
        )
        activate_course(admin_client, event_id, invalid_course_id, admin_courses_csrf)

        fallback_course_name = None
        fallback_course_id = None
        if fallback:
            fallback_course_name = f"{scenario_name}-fallback-course"
            fallback_course_id = create_course(
                admin_client,
                event_id,
                admin_courses_csrf,
                fallback_course_name,
                gender_limit="FEMALE_ONLY",
                teacher_id=int(teacher["id"]),
            )
            activate_course(admin_client, event_id, fallback_course_id, admin_courses_csrf)

        start_round1(admin_client, event_id, admin_courses_csrf)
        prefer_result = submit_round1_preference(target["account"], invalid_course_id, base_url)
        confirm_result = confirm_round1_preference(target["account"], base_url)
        update_student_gender(admin_client, target["student"], admin_students_csrf, "女")
        updated_student = fetch_student_by_id(admin_client, admin_token, int(target["student"]["id"]))
        process_round1(admin_client, event_id, admin_courses_csrf)

        after_round1 = fetch_student_my_selections(target["account"], base_url)
        courses_before_finalize = summarize_courses(fetch_student_courses(target["account"], base_url))
        update_event_times_for_finalize(admin_client, admin_courses_csrf, event_id, scenario_name)
        trigger_snapshot = trigger_finalize(target["account"], base_url)
        after_finalize = fetch_student_my_selections(target["account"], base_url)
        confirmed_course_name = get_confirmed_course_name(after_finalize)
        invalid_after_round1 = pick_selection(after_round1, invalid_course_id)
        invalid_after_finalize = pick_selection(after_finalize, invalid_course_id)

        passed = (
            prefer_result.get("code") == 200
            and confirm_result.get("code") == 200
            and updated_student.get("gender") == "女"
            and invalid_after_round1 is not None
            and invalid_after_round1.get("status") == "CANCELLED"
            and invalid_after_finalize is not None
            and invalid_after_finalize.get("status") == "CANCELLED"
        )
        if fallback:
            passed = passed and confirmed_course_name == fallback_course_name
        else:
            passed = passed and confirmed_course_name is None

        return {
            "eventId": event_id,
            "eventName": scenario_name,
            "invalidCourseId": invalid_course_id,
            "invalidCourseName": invalid_course_name,
            "fallbackCourseId": fallback_course_id,
            "fallbackCourseName": fallback_course_name,
            "preferResult": prefer_result,
            "confirmResult": confirm_result,
            "updatedStudentGender": updated_student.get("gender"),
            "coursesBeforeFinalize": courses_before_finalize,
            "triggerSnapshot": trigger_snapshot,
            "afterRound1Selections": after_round1,
            "afterFinalizeSelections": after_finalize,
            "confirmedCourseName": confirmed_course_name,
            "passed": passed,
        }

    with_fallback = run_single(f"round2-historical-invalid-with-fallback-{datetime.now().strftime('%H%M%S')}", True)
    update_student_gender(
        admin_client,
        fetch_student_by_id(admin_client, admin_token, int(target["student"]["id"])),
        admin_students_csrf,
        "男",
    )
    without_fallback = run_single(f"round2-historical-invalid-no-fallback-{datetime.now().strftime('%H%M%S')}", False)
    return {
        "studentId": target["student"]["id"],
        "studentNo": target["account"].student_no,
        "studentName": target["account"].student_name,
        "scenarioWithFallback": with_fallback,
        "scenarioWithoutFallback": without_fallback,
        "passed": with_fallback["passed"] and without_fallback["passed"],
    }


def run_admin_enroll_guard_scenario(
    admin_client: RemoteClient,
    admin_csrf: str,
    admin_token: str,
    teacher: dict,
    accounts: dict[str, Account],
    base_url: str,
) -> dict:
    male_target = find_student_with_gender_and_account(admin_client, admin_token, accounts, "男")
    event_name = f"Gender-Admin-Enroll-{datetime.now().strftime('%H%M%S')}"
    course_name = f"Admin-Enroll-Guard-{datetime.now().strftime('%H%M%S')}"
    event_id = create_event(admin_client, admin_csrf, event_name, round2_in_future=False)
    save_event_students(admin_client, event_id, [int(male_target["student"]["id"])], admin_csrf)
    course_id = create_course(
        admin_client,
        event_id,
        admin_csrf,
        course_name,
        gender_limit="FEMALE_ONLY",
        teacher_id=int(teacher["id"]),
    )
    activate_course(admin_client, event_id, course_id, admin_csrf)
    page_html, _ = admin_client.get_text(f"/admin/courses/{event_id}/courses/{course_id}/enrollments")
    enroll_csrf = parse_csrf_token(page_html)
    _, final_url = admin_client.post_form_quick(
        f"/admin/courses/{event_id}/courses/{course_id}/enrollments/add",
        [
            ("_csrf", enroll_csrf),
            ("studentId", str(male_target["student"]["id"])),
            ("forceOverflow", "false"),
            ("forceReason", ""),
        ],
    )
    follow_html, _ = admin_client.get_text(final_url.replace(base_url, ""))
    passed = "该课程仅限女生选择" in follow_html
    return {
        "eventId": event_id,
        "courseId": course_id,
        "studentId": male_target["student"]["id"],
        "studentNo": male_target["account"].student_no,
        "redirectUrl": final_url,
        "blockedPromptVisible": passed,
        "passed": passed,
    }


def run_visibility_script(script_path: Path) -> dict:
    result = subprocess.run(
        [sys.executable, str(script_path)],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        check=True,
    )
    summary_path = None
    for line in result.stdout.splitlines():
        line = line.strip()
        if line.endswith(".json") and Path(line).exists():
            summary_path = Path(line)
            break
    if summary_path is None:
        raise RuntimeError("failed to locate visibility summary.json from script output")
    return json.loads(summary_path.read_text(encoding="utf-8"))


def main() -> None:
    args = build_parser().parse_args()
    run_dir = ROOT / args.results_dir / f"gender_limit_suite_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=True)

    admin_client = RemoteClient(args.base_url)
    admin_client.login_web(args.admin_username, args.admin_password)
    admin_courses_page, _ = admin_client.get_text("/admin/courses")
    admin_courses_csrf = parse_csrf_token(admin_courses_page)
    admin_students_page, _ = admin_client.get_text("/admin/students")
    admin_students_csrf = parse_csrf_token(admin_students_page)
    admin_token = admin_client.api_login(args.admin_username, args.admin_password)["data"]["token"]

    accounts = refresh_accounts_from_export(admin_client, run_dir / "student_accounts.xlsx")
    teacher = create_teacher(admin_client, admin_courses_csrf, "gender-limit-suite", args.teacher_password)

    edit_guard = run_edit_guard_scenario(admin_client, admin_courses_csrf, admin_token, accounts, args.base_url)
    pending_filter = run_pending_filter_scenario(
        admin_client,
        admin_courses_csrf,
        admin_students_csrf,
        admin_token,
        accounts,
        args.base_url,
    )
    historical_invalid = run_historical_invalid_scenario(
        admin_client,
        admin_courses_csrf,
        admin_students_csrf,
        admin_token,
        teacher,
        accounts,
        args.base_url,
    )
    admin_enroll_guard = run_admin_enroll_guard_scenario(
        admin_client,
        admin_courses_csrf,
        admin_token,
        teacher,
        accounts,
        args.base_url,
    )
    visibility = run_visibility_script(ROOT / "scripts" / "regression" / "run_gender_reason_visibility_regression.py")

    summary = {
        "teacher": teacher,
        "editGuard": edit_guard,
        "round1PendingFilter": pending_filter,
        "round2HistoricalInvalid": historical_invalid,
        "adminEnrollGuard": admin_enroll_guard,
        "visibilityRegression": visibility,
        "passed": all(
            [
                edit_guard["passed"],
                pending_filter["passed"],
                historical_invalid["passed"],
                admin_enroll_guard["passed"],
                visibility["passed"],
            ]
        ),
    }
    summary_path = run_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(summary_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
