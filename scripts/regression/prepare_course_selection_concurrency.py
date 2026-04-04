import argparse
import html
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta
from http.cookiejar import CookieJar
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from course_selection_regression_lib import DEFAULT_COURSE_SPECS, env_or_required


ADMIN_CLASS_TYPE = "行政班"


def build_arg_parser() -> argparse.ArgumentParser:
    base_url, base_url_required = env_or_required("COURSE_SELECTION_BASE_URL")
    admin_username, admin_username_required = env_or_required("COURSE_SELECTION_ADMIN_USERNAME")
    admin_password, admin_password_required = env_or_required("COURSE_SELECTION_ADMIN_PASSWORD")

    parser = argparse.ArgumentParser(description="准备远程抢课并发测试数据。")
    parser.add_argument("--base-url", default=base_url, required=base_url_required)
    parser.add_argument("--admin-username", default=admin_username, required=admin_username_required)
    parser.add_argument("--admin-password", default=admin_password, required=admin_password_required)
    parser.add_argument("--school-id", default="")
    parser.add_argument("--grade-name", default="高一")
    parser.add_argument("--class-count", type=int, default=10)
    parser.add_argument("--students-per-class", type=int, default=50)
    parser.add_argument("--prefix", default="")
    parser.add_argument("--output-dir", default="scripts/regression/generated")
    parser.add_argument("--event-id", type=int, default=0)
    return parser


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def strip_tags(value: str) -> str:
    no_tags = re.sub(r"<[^>]+>", "", value or "", flags=re.S)
    return html.unescape(no_tags).strip()


def parse_csrf_token(html_text: str) -> str:
    patterns = [
        r'name="_csrf"[^>]*value="([^"]+)"',
        r'value="([^"]+)"[^>]*name="_csrf"',
    ]
    for pattern in patterns:
        match = re.search(pattern, html_text)
        if match:
            return html.unescape(match.group(1))
    raise RuntimeError("未能从页面中提取 CSRF token")


def extract_alert(html_text: str) -> Dict[str, str]:
    result = {}
    success = re.search(r'<div[^>]*class="alert alert-success"[^>]*>(.*?)</div>', html_text, re.S)
    error = re.search(r'<div[^>]*class="alert alert-error"[^>]*>(.*?)</div>', html_text, re.S)
    if success:
        result["success"] = strip_tags(success.group(1))
    if error:
        result["error"] = strip_tags(error.group(1))
    return result


def make_multipart(fields: Dict[str, str], files: Dict[str, Tuple[str, bytes, str]]) -> Tuple[bytes, str]:
    boundary = f"----CodexBoundary{int(time.time() * 1000)}"
    chunks: List[bytes] = []
    for key, value in fields.items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                f'Content-Disposition: form-data; name="{key}"\r\n\r\n'.encode(),
                str(value).encode("utf-8"),
                b"\r\n",
            ]
        )
    for key, (filename, data, content_type) in files.items():
        chunks.extend(
            [
                f"--{boundary}\r\n".encode(),
                (
                    f'Content-Disposition: form-data; name="{key}"; filename="{filename}"\r\n'
                    f"Content-Type: {content_type}\r\n\r\n"
                ).encode(),
                data,
                b"\r\n",
            ]
        )
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


class RemoteClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.cookie_jar = CookieJar()
        self.web_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(self.cookie_jar))

    def _full_url(self, path: str) -> str:
        if path.startswith("http://") or path.startswith("https://"):
            return path
        if not path.startswith("/"):
            path = "/" + path
        return self.base_url + path

    def get_text(self, path: str, headers: Optional[Dict[str, str]] = None):
        req = urllib.request.Request(self._full_url(path), headers=headers or {})
        with self.web_opener.open(req, timeout=30) as resp:
            return resp.read().decode("utf-8", errors="replace"), resp.geturl()

    def get_bytes(self, path: str, headers: Optional[Dict[str, str]] = None):
        req = urllib.request.Request(self._full_url(path), headers=headers or {})
        with self.web_opener.open(req, timeout=60) as resp:
            return resp.read(), resp.geturl(), {k: v for k, v in resp.headers.items()}

    def post_form(self, path: str, fields, headers: Optional[Dict[str, str]] = None):
        encoded = urllib.parse.urlencode(list(fields)).encode("utf-8")
        req_headers = {"Content-Type": "application/x-www-form-urlencoded"}
        if headers:
            req_headers.update(headers)
        req = urllib.request.Request(self._full_url(path), data=encoded, headers=req_headers, method="POST")
        with self.web_opener.open(req, timeout=60) as resp:
            return resp.read().decode("utf-8", errors="replace"), resp.geturl()

    def post_multipart(self, path: str, fields: Dict[str, str], files: Dict[str, Tuple[str, bytes, str]]):
        body, content_type = make_multipart(fields, files)
        req = urllib.request.Request(
            self._full_url(path),
            data=body,
            headers={"Content-Type": content_type},
            method="POST",
        )
        with self.web_opener.open(req, timeout=120) as resp:
            return resp.read().decode("utf-8", errors="replace"), resp.geturl()

    def login_web(self, username: str, password: str) -> None:
        login_html, _ = self.get_text("/login")
        csrf = parse_csrf_token(login_html)
        response_html, final_url = self.post_form("/login", [("_csrf", csrf), ("username", username), ("password", password)])
        if final_url.endswith("/login") or "/login?" in final_url:
            alerts = extract_alert(response_html)
            raise RuntimeError(f"网页登录失败：{alerts.get('error') or final_url}")

    def api_request(self, method: str, path: str, token: Optional[str] = None, json_body: Optional[dict] = None):
        headers = {}
        data = None
        if token:
            headers["Authorization"] = f"Bearer {token}"
        if json_body is not None:
            headers["Content-Type"] = "application/json"
            data = json.dumps(json_body, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(self._full_url(path), data=data, headers=headers, method=method.upper())
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))

    def api_login(self, username: str, password: str) -> str:
        payload = self.api_request("POST", "/api/auth/login", json_body={"username": username, "password": password})
        token = (payload.get("data") or {}).get("token")
        if not token:
            raise RuntimeError(f"API 登录失败：{payload}")
        return token


def generate_workbooks(output_dir: Path, prefix: str, grade_name: str, class_count: int, students_per_class: int):
    ensure_dir(output_dir)
    classes_path = output_dir / f"{prefix}_classes.xlsx"
    students_path = output_dir / f"{prefix}_students.xlsx"

    wb_classes = Workbook()
    ws_classes = wb_classes.active
    ws_classes.title = "班级"
    ws_classes.append(["班级名称", "类型", "年级"])
    for class_index in range(1, class_count + 1):
        ws_classes.append([f"{class_index}班", ADMIN_CLASS_TYPE, grade_name])
    wb_classes.save(classes_path)

    wb_students = Workbook()
    ws_students = wb_students.active
    ws_students.title = "学生"
    ws_students.append(["年级", "班级", "姓名", "性别", "学号", "身份证号", "选修课", "学籍状态"])
    sequence = 1
    for class_index in range(1, class_count + 1):
        class_name = f"{class_index}班"
        for student_index in range(1, students_per_class + 1):
            name = f"{prefix}-{class_index:02d}-{student_index:03d}"
            gender = "男" if student_index % 2 == 1 else "女"
            student_no = f"3{datetime.now().strftime('%y%m%d')}{sequence:04d}"
            id_card = f"11010120080101{sequence:04d}"
            ws_students.append([grade_name, class_name, name, gender, student_no, id_card, "", "在籍"])
            sequence += 1
    wb_students.save(students_path)
    return classes_path, students_path


def import_file(client: RemoteClient, path: str, file_path: Path) -> Dict[str, str]:
    page_html, _ = client.get_text("/admin/import")
    csrf = parse_csrf_token(page_html)
    html_text, _ = client.post_multipart(
        path,
        {"_csrf": csrf},
        {"file": (file_path.name, file_path.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    return extract_alert(html_text)


def fetch_classes(client: RemoteClient, token: str, grade_name: str, class_count: int) -> List[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=token)
    data = payload.get("data") or []
    target_names = {f"{index}班" for index in range(1, class_count + 1)}
    result = []
    for item in data:
        if item.get("gradeName") == grade_name and item.get("name") in target_names and item.get("type") == ADMIN_CLASS_TYPE:
            result.append(item)
    if len(result) != class_count:
        raise RuntimeError(f"仅找到 {len(result)} 个目标行政班，预期 {class_count} 个")
    result.sort(key=lambda item: int(str(item["name"]).replace("班", "")))
    return result


def fetch_students_from_classes(client: RemoteClient, token: str, class_rows: List[dict], prefix: str, students_per_class: int):
    collected = []
    for class_row in class_rows:
        path = f"/api/admin/students?classId={class_row['id']}&page=0&size={students_per_class + 20}"
        payload = client.api_request("GET", path, token=token)
        content = (payload.get("data") or {}).get("content") or []
        matched = [item for item in content if str(item.get("name", "")).startswith(prefix)]
        if len(matched) != students_per_class:
            raise RuntimeError(f"班级 {class_row['name']} 仅匹配到 {len(matched)} 名测试学生，预期 {students_per_class} 名")
        collected.extend(matched)
    return collected


def create_event(client: RemoteClient, event_name: str) -> int:
    page_html, _ = client.get_text("/admin/courses")
    csrf = parse_csrf_token(page_html)
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(minutes=10)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now + timedelta(days=3)).strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"未能从跳转链接解析活动 ID：{final_url}")
    return int(match.group(1))


def save_event_students(client: RemoteClient, event_id: int, student_ids: List[int]) -> Dict[str, str]:
    detail_html, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=students")
    csrf = parse_csrf_token(detail_html)
    fields = [("_csrf", csrf)]
    for student_id in student_ids:
        fields.append(("studentIds", str(student_id)))
    html_text, _ = client.post_form(f"/admin/courses/{event_id}/students/save", fields)
    return extract_alert(html_text)


def export_student_accounts(client: RemoteClient, output_path: Path, student_ids: List[int]) -> Path:
    ensure_dir(output_path.parent)
    params = "&".join(f"studentIds={student_id}" for student_id in student_ids)
    try:
        data, _, headers = client.get_bytes(f"/admin/student-accounts/export?{params}")
    except urllib.error.HTTPError as exc:
        if exc.code != 400:
            raise
        data, _, headers = client.get_bytes("/admin/student-accounts/export")
    content_type = headers.get("Content-Type", "")
    if "spreadsheetml" not in content_type:
        raise RuntimeError("导出学生账号失败：未返回 xlsx 文件")
    output_path.write_bytes(data)
    return output_path


def create_courses(client: RemoteClient, event_id: int, class_rows: List[dict]) -> Dict[str, Dict[str, object]]:
    detail_html, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    csrf = parse_csrf_token(detail_html)
    class_ids = [str(item["id"]) for item in class_rows]
    for name, mode, capacity in DEFAULT_COURSE_SPECS:
        fields = [
            ("_csrf", csrf),
            ("name", name),
            ("description", f"{name} 并发回归测试课程"),
            ("capacityMode", mode),
        ]
        if mode == "GLOBAL":
            fields.append(("totalCapacity", str(capacity)))
        else:
            for class_id in class_ids:
                fields.append(("classIds", class_id))
                fields.append(("classCapacities", str(capacity)))
        client.post_form(f"/admin/courses/{event_id}/courses/save", fields)

    html_text, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    return parse_course_rows(html_text, event_id)


def parse_course_rows(html_text: str, event_id: int) -> Dict[str, Dict[str, object]]:
    pattern = re.compile(
        rf"<tr[^>]*>\s*<td[^>]*>(?P<name>.*?)</td>.*?/admin/courses/{event_id}/courses/(?P<course_id>\d+)/enrollments",
        re.S,
    )
    result = {}
    for match in pattern.finditer(html_text):
        result[strip_tags(match.group("name"))] = {"id": int(match.group("course_id"))}
    return result


def activate_courses(client: RemoteClient, event_id: int, course_rows: Dict[str, Dict[str, object]]) -> None:
    detail_html, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    csrf = parse_csrf_token(detail_html)
    for course in course_rows.values():
        client.post_form(f"/admin/courses/{event_id}/courses/{course['id']}/activate", [("_csrf", csrf)])


def start_and_process_round2(client: RemoteClient, event_id: int) -> str:
    detail_html, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    csrf = parse_csrf_token(detail_html)
    client.post_form(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])

    detail_html, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    csrf = parse_csrf_token(detail_html)
    client.post_form(f"/admin/courses/events/{event_id}/process", [("_csrf", csrf)])

    for _ in range(60):
        payload, _ = client.get_text(f"/admin/courses/events/{event_id}/lottery-status")
        status = json.loads(payload).get("status")
        if status and status != "PROCESSING":
            return status
        time.sleep(2)
    raise RuntimeError("等待抽签结算进入 ROUND2 超时")


def read_first_account(export_path: Path) -> Dict[str, str]:
    workbook = load_workbook(export_path)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    index_map = {str(value): idx for idx, value in enumerate(headers)}
    required = ["姓名", "学号", "登录账号", "初始密码"]
    for column in required:
        if column not in index_map:
            raise RuntimeError(f"导出账号文件缺少列：{column}")
    row = sheet[2]
    return {
        "name": row[index_map["姓名"]].value,
        "studentNo": row[index_map["学号"]].value,
        "loginId": row[index_map["登录账号"]].value,
        "password": row[index_map["初始密码"]].value,
    }


def verify_student_access(client: RemoteClient, account: Dict[str, str]) -> Dict[str, object]:
    token = client.api_request("POST", "/api/auth/login", json_body={"username": account["loginId"], "password": account["password"]})["data"]["token"]
    event_payload = client.api_request("GET", "/api/student/events/current", token=token)
    courses_payload = client.api_request("GET", "/api/student/courses", token=token)
    return {"event": event_payload.get("data"), "courseCount": len(courses_payload.get("data") or [])}


def main() -> int:
    args = build_arg_parser().parse_args()
    prefix = args.prefix or f"QC{datetime.now().strftime('%m%d')}"
    output_dir = Path(args.output_dir)
    event_name = f"{datetime.now().strftime('%Y-%m-%d')} {args.grade_name} 抢课并发测试 {prefix}"

    classes_path, students_path = generate_workbooks(
        output_dir=output_dir,
        prefix=prefix,
        grade_name=args.grade_name,
        class_count=args.class_count,
        students_per_class=args.students_per_class,
    )

    client = RemoteClient(args.base_url)
    client.login_web(args.admin_username, args.admin_password)
    admin_token = client.api_login(args.admin_username, args.admin_password)

    class_import_result = import_file(client, "/admin/import/classes", classes_path)
    student_import_result = import_file(client, "/admin/import/students", students_path)
    class_rows = fetch_classes(client, admin_token, args.grade_name, args.class_count)
    students = fetch_students_from_classes(client, admin_token, class_rows, prefix, args.students_per_class)
    student_ids = [int(item["id"]) for item in students]

    event_id = args.event_id or create_event(client, event_name)
    event_student_result = save_event_students(client, event_id, student_ids)
    accounts_path = output_dir / f"{prefix}_student_accounts.xlsx"
    export_student_accounts(client, accounts_path, student_ids)

    course_rows = create_courses(client, event_id, class_rows)
    activate_courses(client, event_id, course_rows)
    final_status = start_and_process_round2(client, event_id)

    sample_account = read_first_account(accounts_path)
    sample_verification = verify_student_access(client, sample_account)

    summary = {
        "baseUrl": args.base_url,
        "schoolId": args.school_id,
        "eventId": event_id,
        "eventName": event_name,
        "eventStatus": final_status,
        "prefix": prefix,
        "classImport": class_import_result,
        "studentImport": student_import_result,
        "eventStudents": event_student_result,
        "studentCount": len(student_ids),
        "classCount": len(class_rows),
        "courses": {name: value["id"] for name, value in sorted(course_rows.items())},
        "files": {
            "classes": str(classes_path),
            "students": str(students_path),
            "accounts": str(accounts_path),
        },
        "sampleAccount": sample_account,
        "sampleVerification": sample_verification,
    }

    summary_path = output_dir / f"{prefix}_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    print(f"SUMMARY_FILE={summary_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        print(body)
        raise
