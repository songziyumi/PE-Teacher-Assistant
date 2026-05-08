import argparse
import json
import re
import sys
import time
import urllib.parse
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "regression"))

from course_selection_regression_lib import (  # noqa: E402
    Account,
    RemoteClient,
    api_login_student,
    parse_csrf_token,
    reset_student_passwords,
)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Gender-limit round2 auto-assignment regression")
    parser.add_argument("--base-url", default="http://127.0.0.1:8080")
    parser.add_argument("--admin-username", default="admin")
    parser.add_argument("--admin-password", default="Admin@2024!")
    parser.add_argument(
        "--accounts-csv",
        default="scripts/regression/results/prepare_qc0404r4/QC0404R4_student_accounts.csv",
    )
    parser.add_argument("--results-dir", default="temp/test-results")
    parser.add_argument("--prefix", default="QC0404R4")
    parser.add_argument("--teacher-password", default="abc127!!!")
    return parser


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def load_accounts(path: Path) -> list[Account]:
    import csv

    rows: list[Account] = []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            rows.append(
                Account(
                    username=row["username"].strip(),
                    password=row["password"].strip(),
                    student_no=row["student_no"].strip(),
                    student_name=row["student_name"].strip(),
                )
            )
    return rows


def close_existing_events(client: RemoteClient, csrf: str) -> list[int]:
    page, _ = client.get_text("/admin/courses")
    event_ids = sorted({int(match) for match in re.findall(r"/admin/courses/events/(\d+)/close", page)})
    for event_id in event_ids:
        client.post_form_quick(f"/admin/courses/events/{event_id}/close", [("_csrf", csrf)])
    return event_ids


def create_teacher(client: RemoteClient, csrf: str, prefix: str, password: str) -> dict:
    phone = f"19{int(time.time() * 1000) % 1_000_000_000:09d}"
    name = f"{prefix}-GenderTeacher-{phone[-4:]}"
    client.post_form_quick(
        "/admin/teachers/add",
        [
            ("_csrf", csrf),
            ("name", name),
            ("password", password),
            ("phone", phone),
            ("accountType", "TEACHER"),
        ],
    )
    page, _ = client.get_text(f"/admin/teachers?phone={urllib.parse.quote_plus(phone)}")
    match = re.search(r"/admin/teachers/reset-password/(\d+)", page)
    if not match:
        raise RuntimeError(f"failed to create teacher: {phone}")
    return {"id": int(match.group(1)), "username": phone, "password": password, "name": name}


def fetch_all_classes(client: RemoteClient, admin_token: str) -> list[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=admin_token)
    return payload.get("data", [])


def fetch_students_by_prefix(client: RemoteClient, admin_token: str, classes: list[dict], prefix: str) -> list[dict]:
    matches: list[dict] = []
    seen_ids: set[int] = set()
    for class_row in classes:
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            name = str(item.get("name", ""))
            student_id = int(item["id"])
            if name.startswith(prefix) and student_id not in seen_ids:
                seen_ids.add(student_id)
                matches.append(item)
    return matches


def pick_gender_students(accounts: list[Account], students: list[dict]) -> tuple[dict, dict, dict]:
    account_by_no = {item.student_no: item for item in accounts}
    male: list[dict] = []
    female: list[dict] = []
    for student in students:
        gender = str(student.get("gender", "")).strip()
        student_no = str(student.get("studentNo", "")).strip()
        if student_no not in account_by_no:
            continue
        merged = {"student": student, "account": account_by_no[student_no]}
        normalized_gender = gender.lower()
        if gender == "男" or normalized_gender == "male":
            male.append(merged)
            continue
        if gender == "女" or normalized_gender == "female":
            female.append(merged)
            continue
        if gender == "男":
            male.append(merged)
        elif gender == "女":
            female.append(merged)
    if len(male) < 2 or len(female) < 1:
        raise RuntimeError("not enough male/female test students found")
    return male[0], female[0], male[1]


def refresh_accounts_from_export(client: RemoteClient, output_path: Path) -> dict[str, Account]:
    from openpyxl import load_workbook

    output_path.write_bytes(client.get_bytes("/admin/student-accounts/export"))
    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    accounts: dict[str, Account] = {}
    for row in rows[1:]:
        if not row or not row[1] or not row[4] or not row[6]:
            continue
        account = Account(
            username=str(row[4]).strip(),
            password=str(row[6]).strip(),
            student_no=str(row[1]).strip(),
            student_name=str(row[0]).strip(),
        )
        accounts[account.student_no] = account
    return accounts


def create_event(client: RemoteClient, csrf: str, event_name: str) -> int:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now + timedelta(minutes=5)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now + timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form_quick("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"failed to create event: {final_url}")
    return int(match.group(1))


def update_event_round2_end(client: RemoteClient, csrf: str, event_id: int, event_name: str, end_time: datetime) -> None:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("id", str(event_id)),
        ("name", event_name),
        ("round1Start", (now - timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(minutes=30)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", end_time.strftime("%Y-%m-%dT%H:%M")),
    ]
    client.post_form_quick("/admin/courses/events/save", fields)


def save_event_students(client: RemoteClient, event_id: int, student_ids: list[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick(f"/admin/courses/{event_id}/students/save", fields)


def create_gender_course(
    client: RemoteClient,
    event_id: int,
    csrf: str,
    name: str,
    teacher_id: int,
    gender_limit: str,
    capacity: int,
) -> None:
    client.post_form_quick(
        f"/admin/courses/{event_id}/courses/save",
        [
            ("_csrf", csrf),
            ("name", name),
            ("description", f"{name} gender-limit regression"),
            ("teacherId", str(teacher_id)),
            ("capacityMode", "GLOBAL"),
            ("genderLimit", gender_limit),
            ("totalCapacity", str(capacity)),
        ],
    )


def parse_course_ids(page: str) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for match in re.finditer(r'data-course-id="(?P<id>\d+)"[^>]*data-course-name="(?P<name>[^"]+)"', page):
        mapping[match.group("name")] = int(match.group("id"))
    return mapping


def activate_all_courses(client: RemoteClient, event_id: int, csrf: str) -> None:
    page, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    course_ids = [int(match) for match in re.findall(rf"/admin/courses/{event_id}/courses/(\d+)/activate", page)]
    for course_id in course_ids:
        client.post_form_quick(f"/admin/courses/{event_id}/courses/{course_id}/activate", [("_csrf", csrf)])


def start_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])


def process_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/process", [("_csrf", csrf)])
    for _ in range(180):
        payload, final_url = client.get_text(f"/admin/courses/events/{event_id}/lottery-status")
        if "/login" in final_url:
            time.sleep(1)
            continue
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            time.sleep(1)
            continue
        if data.get("status") and data.get("status") != "PROCESSING":
            return
        time.sleep(1)
    raise RuntimeError("round1 processing timeout")


def fetch_student_courses(account: Account, base_url: str) -> list[dict]:
    client, login = api_login_student(account, base_url)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/courses", token=token)
    return payload.get("data", [])


def fetch_student_event_current(account: Account, base_url: str) -> dict:
    client, login = api_login_student(account, base_url)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/events/current", token=token)
    return payload.get("data")


def fetch_student_my_selections(account: Account, base_url: str) -> list[dict]:
    client, login = api_login_student(account, base_url)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/my-selections", token=token)
    return payload.get("data", [])


def get_confirmed_course_name(selections: list[dict]) -> str | None:
    for selection in selections:
        if selection.get("status") == "CONFIRMED":
            return selection.get("courseName")
    return None


def login_teacher(base_url: str, username: str, password: str) -> RemoteClient:
    client = RemoteClient(base_url)
    client.login_web(username, password)
    return client


def activate_student_password(account: Account, base_url: str) -> Account:
    client = RemoteClient(base_url)
    client.login_web(account.username, account.password)
    page_html, final_url = client.get_text("/student/password?force=true")
    if "force=true" not in final_url:
        return account
    csrf = parse_csrf_token(page_html)
    new_password = f"{account.password}A1"
    login_alias = f"GL{account.student_no[-8:]}"
    client.post_form(
        "/student/password",
        [
            ("_csrf", csrf),
            ("force", "true"),
            ("oldPassword", account.password),
            ("newPassword", new_password),
            ("confirmPassword", new_password),
            ("loginAlias", login_alias),
        ],
    )
    account.password = new_password
    return account


def assert_page_ok(page_html: str, final_url: str, marker: str) -> dict:
    return {
        "ok": final_url.endswith(marker) or marker in final_url,
        "finalUrl": final_url,
        "contains500": (
            "Whitelabel Error Page" in page_html
            or "Internal Server Error" in page_html
            or "<title>500" in page_html
            or ">500<" in page_html
        ),
    }


def run_assignment_scenario(
    admin_client: RemoteClient,
    csrf: str,
    base_url: str,
    teacher: dict,
    event_name: str,
    participants: list[dict],
    course_specs: list[dict],
) -> dict:
    event_id = create_event(admin_client, csrf, event_name)
    save_event_students(admin_client, event_id, [int(item["student"]["id"]) for item in participants], csrf)
    for spec in course_specs:
        create_gender_course(
            admin_client,
            event_id,
            csrf,
            spec["name"],
            int(teacher["id"]),
            spec["genderLimit"],
            int(spec["capacity"]),
        )
    activate_all_courses(admin_client, event_id, csrf)
    detail_page, _ = admin_client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    course_ids = parse_course_ids(detail_page)
    start_round1(admin_client, event_id, csrf)
    process_round1(admin_client, event_id, csrf)

    before_course_views = {}
    for item in participants:
        before_course_views[item["account"].student_no] = fetch_student_courses(item["account"], base_url)

    update_event_round2_end(admin_client, csrf, event_id, event_name, datetime.now() - timedelta(minutes=1))
    trigger_snapshot = fetch_student_event_current(participants[0]["account"], base_url)

    selections_after = {}
    for item in participants:
        selections_after[item["account"].student_no] = fetch_student_my_selections(item["account"], base_url)

    return {
        "eventId": event_id,
        "eventName": event_name,
        "courseIds": course_ids,
        "beforeCourses": before_course_views,
        "triggerSnapshot": trigger_snapshot,
        "afterSelections": selections_after,
    }


def summarize_course_visibility(courses: list[dict]) -> dict[str, dict]:
    result = {}
    for item in courses:
        result[str(item["name"])] = {
            "eligible": item.get("eligible"),
            "genderLimit": item.get("genderLimit"),
            "genderLimitLabel": item.get("genderLimitLabel"),
            "ineligibleMessage": item.get("ineligibleMessage"),
            "remaining": item.get("remaining"),
        }
    return result


def main() -> None:
    args = build_parser().parse_args()
    results_dir = ROOT / args.results_dir
    ensure_dir(results_dir)
    run_dir = results_dir / f"gender_limit_round2_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    ensure_dir(run_dir)

    accounts = load_accounts(ROOT / args.accounts_csv)
    admin_client = RemoteClient(args.base_url)
    admin_client.login_web(args.admin_username, args.admin_password)
    admin_page, _ = admin_client.get_text("/admin/courses")
    admin_csrf = parse_csrf_token(admin_page)
    admin_token = admin_client.api_login(args.admin_username, args.admin_password)["data"]["token"]

    closed_existing = close_existing_events(admin_client, admin_csrf)
    admin_page, _ = admin_client.get_text("/admin/courses")
    admin_csrf = parse_csrf_token(admin_page)

    export_accounts = refresh_accounts_from_export(admin_client, run_dir / "accounts_reset.xlsx")
    teacher = create_teacher(admin_client, admin_csrf, args.prefix, args.teacher_password)
    classes = fetch_all_classes(admin_client, admin_token)
    students = fetch_students_by_prefix(admin_client, admin_token, classes, args.prefix)
    try:
        male1, female1, male2 = pick_gender_students(accounts, students)
    except RuntimeError:
        students = fetch_students_by_prefix(admin_client, admin_token, classes, "")
        male1, female1, male2 = pick_gender_students(list(export_accounts.values()), students)

    reset_student_passwords(
        admin_client,
        [int(male1["student"]["id"]), int(female1["student"]["id"]), int(male2["student"]["id"])],
        admin_csrf,
    )
    export_accounts = refresh_accounts_from_export(admin_client, run_dir / "accounts_reset.xlsx")
    for target in (male1, female1, male2):
        student_no = target["account"].student_no
        if student_no in export_accounts:
            target["account"] = export_accounts[student_no]

    scenario1 = run_assignment_scenario(
        admin_client,
        admin_csrf,
        args.base_url,
        teacher,
        f"{args.prefix}-gender-round2-match-{datetime.now().strftime('%H%M%S')}",
        [male1, female1],
        [
            {"name": "Gender-A-Girls", "genderLimit": "FEMALE_ONLY", "capacity": 5},
            {"name": "Gender-B-Boys", "genderLimit": "MALE_ONLY", "capacity": 5},
        ],
    )

    male1_before = summarize_course_visibility(scenario1["beforeCourses"][male1["account"].student_no])
    female1_before = summarize_course_visibility(scenario1["beforeCourses"][female1["account"].student_no])
    male1_confirmed = get_confirmed_course_name(scenario1["afterSelections"][male1["account"].student_no])
    female1_confirmed = get_confirmed_course_name(scenario1["afterSelections"][female1["account"].student_no])

    scenario2 = run_assignment_scenario(
        admin_client,
        admin_csrf,
        args.base_url,
        teacher,
        f"{args.prefix}-gender-round2-none-{datetime.now().strftime('%H%M%S')}",
        [male2],
        [
            {"name": "Gender-C-Girls-Only", "genderLimit": "FEMALE_ONLY", "capacity": 5},
        ],
    )
    male2_before = summarize_course_visibility(scenario2["beforeCourses"][male2["account"].student_no])
    male2_confirmed = get_confirmed_course_name(scenario2["afterSelections"][male2["account"].student_no])

    admin_enrollment_checks = {}
    for course_name, course_id in scenario1["courseIds"].items():
        page_html, final_url = admin_client.get_text(f"/admin/courses/{scenario1['eventId']}/courses/{course_id}/enrollments")
        admin_enrollment_checks[course_name] = assert_page_ok(
            page_html,
            final_url,
            f"/admin/courses/{scenario1['eventId']}/courses/{course_id}/enrollments",
        )
    for course_name, course_id in scenario2["courseIds"].items():
        page_html, final_url = admin_client.get_text(f"/admin/courses/{scenario2['eventId']}/courses/{course_id}/enrollments")
        admin_enrollment_checks[course_name] = assert_page_ok(
            page_html,
            final_url,
            f"/admin/courses/{scenario2['eventId']}/courses/{course_id}/enrollments",
        )

    teacher_client = login_teacher(args.base_url, teacher["username"], teacher["password"])
    teacher_enrollment_checks = {}
    for course_name, course_id in scenario1["courseIds"].items():
        page_html, final_url = teacher_client.get_text(f"/teacher/courses/{scenario1['eventId']}/courses/{course_id}/enrollments")
        teacher_enrollment_checks[course_name] = assert_page_ok(
            page_html,
            final_url,
            f"/teacher/courses/{scenario1['eventId']}/courses/{course_id}/enrollments",
        )
    for course_name, course_id in scenario2["courseIds"].items():
        page_html, final_url = teacher_client.get_text(f"/teacher/courses/{scenario2['eventId']}/courses/{course_id}/enrollments")
        teacher_enrollment_checks[course_name] = assert_page_ok(
            page_html,
            final_url,
            f"/teacher/courses/{scenario2['eventId']}/courses/{course_id}/enrollments",
        )

    student_page_checks = {}
    for target in (male1, female1, male2):
        target["account"] = activate_student_password(target["account"], args.base_url)
        client = RemoteClient(args.base_url)
        client.login_web(target["account"].username, target["account"].password)
        courses_html, courses_url = client.get_text("/student/courses")
        my_courses_html, my_courses_url = client.get_text("/student/my-courses")
        student_page_checks[target["account"].student_no] = {
            "coursesPage": assert_page_ok(courses_html, courses_url, "/student/courses"),
            "myCoursesPage": assert_page_ok(my_courses_html, my_courses_url, "/student/my-courses"),
        }

    summary = {
        "closedExistingEventIds": closed_existing,
        "teacher": teacher,
        "participants": {
            "male1": {
                "studentId": male1["student"]["id"],
                "studentNo": male1["account"].student_no,
                "studentName": male1["account"].student_name,
                "gender": male1["student"].get("gender"),
            },
            "female1": {
                "studentId": female1["student"]["id"],
                "studentNo": female1["account"].student_no,
                "studentName": female1["account"].student_name,
                "gender": female1["student"].get("gender"),
            },
            "male2": {
                "studentId": male2["student"]["id"],
                "studentNo": male2["account"].student_no,
                "studentName": male2["account"].student_name,
                "gender": male2["student"].get("gender"),
            },
        },
        "scenario1": {
            "eventId": scenario1["eventId"],
            "eventName": scenario1["eventName"],
            "male1BeforeCourses": male1_before,
            "female1BeforeCourses": female1_before,
            "male1ConfirmedCourse": male1_confirmed,
            "female1ConfirmedCourse": female1_confirmed,
            "rawAfterSelections": scenario1["afterSelections"],
            "passed": male1_confirmed == "Gender-B-Boys" and female1_confirmed == "Gender-A-Girls",
        },
        "scenario2": {
            "eventId": scenario2["eventId"],
            "eventName": scenario2["eventName"],
            "male2BeforeCourses": male2_before,
            "male2ConfirmedCourse": male2_confirmed,
            "rawAfterSelections": scenario2["afterSelections"],
            "passed": male2_confirmed is None,
        },
        "adminEnrollmentChecks": admin_enrollment_checks,
        "teacherEnrollmentChecks": teacher_enrollment_checks,
        "studentPageChecks": student_page_checks,
    }

    summary_path = run_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(summary_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
