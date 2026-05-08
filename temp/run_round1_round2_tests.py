import csv
import html
import json
import math
import re
import statistics
import threading
import time
import urllib.parse
import urllib.request
from urllib.error import HTTPError
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime, timedelta
from http.cookiejar import CookieJar
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook


BASE_URL = "http://175.24.131.74:8080"
ADMIN_USERNAME = "qmadmin"
ADMIN_PASSWORD = "abc127!!!"
ACCOUNTS_CSV = Path("temp/jmeter/qc0403_accounts.csv")
RESULT_DIR = Path("temp/test-results")
GRADE_NAME = "高一"
CLASS_NAMES = [f"{index}班" for index in range(1, 11)]
COURSE_SPECS = [
    ("篮球", "GLOBAL", 120),
    ("足球", "GLOBAL", 40),
    ("排球", "GLOBAL", 40),
    ("羽毛球", "PER_CLASS", 4),
    ("飞盘", "GLOBAL", 40),
    ("匹克球", "GLOBAL", 40),
    ("手球", "GLOBAL", 40),
    ("武术", "GLOBAL", 40),
    ("啦啦操", "GLOBAL", 40),
    ("乒乓球", "GLOBAL", 80),
]
COURSE_CAPACITY = {name: capacity for name, _, capacity in COURSE_SPECS}
ROUND1_PREF1 = "篮球"
ROUND1_PREF2 = "乒乓球"
ROUND2_TARGET = "足球"
PREFERENCE_WORKERS = 40
RESULT_WORKERS = 50
ROUND2_MAX_USERS = 200
CONFIRM_WORKERS = 5


def ensure_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)


def parse_csrf_token(page: str) -> str:
    patterns = [
        r'name="_csrf"[^>]*value="([^"]+)"',
        r'value="([^"]+)"[^>]*name="_csrf"',
    ]
    for pattern in patterns:
        match = re.search(pattern, page)
        if match:
            return html.unescape(match.group(1))
    raise RuntimeError("未找到 CSRF token")


def percentile(values: List[float], ratio: float) -> float:
    if not values:
        return 0.0
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    index = (len(ordered) - 1) * ratio
    lower = math.floor(index)
    upper = math.ceil(index)
    if lower == upper:
        return ordered[lower]
    return ordered[lower] + (ordered[upper] - ordered[lower]) * (index - lower)


@dataclass
class Account:
    username: str
    password: str
    student_no: str
    student_name: str


def next_password(old_password: str) -> str:
    return f"{old_password}A1"


class RemoteClient:
    def __init__(self, base_url: str):
        self.base_url = base_url.rstrip("/")
        self.cookie_jar = CookieJar()
        self.web_opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(self.cookie_jar))
        self.quick_opener = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar),
            NoRedirectHandler(),
        )

    def _url(self, path: str) -> str:
        if path.startswith("http://") or path.startswith("https://"):
            return path
        if not path.startswith("/"):
            path = "/" + path
        return self.base_url + path

    def _open(self, opener, request, timeout: int):
        last_error = None
        for _ in range(3):
            try:
                if hasattr(opener, "open"):
                    return opener.open(request, timeout=timeout)
                return opener(request, timeout=timeout)
            except Exception as exc:
                last_error = exc
                time.sleep(1)
        raise last_error

    def get_text(self, path: str) -> Tuple[str, str]:
        request = urllib.request.Request(self._url(path))
        with self._open(self.web_opener, request, 180) as response:
            return response.read().decode("utf-8", errors="replace"), response.geturl()

    def get_bytes(self, path: str) -> bytes:
        request = urllib.request.Request(self._url(path))
        with self._open(self.web_opener, request, 180) as response:
            return response.read()

    def post_form(self, path: str, fields: List[Tuple[str, str]]) -> Tuple[str, str]:
        payload = urllib.parse.urlencode(fields).encode("utf-8")
        request = urllib.request.Request(
            self._url(path),
            data=payload,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )
        with self._open(self.web_opener, request, 180) as response:
            return response.read().decode("utf-8", errors="replace"), response.geturl()

    def post_form_quick(self, path: str, fields: List[Tuple[str, str]]) -> Tuple[int, str]:
        payload = urllib.parse.urlencode(fields).encode("utf-8")
        request = urllib.request.Request(
            self._url(path),
            data=payload,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            method="POST",
        )
        with self._open(self.quick_opener, request, 180) as response:
            location = response.headers.get("Location", "") or response.geturl()
            response.read()
            return response.getcode(), location

    def login_web(self, username: str, password: str) -> None:
        page, _ = self.get_text("/login")
        csrf = parse_csrf_token(page)
        _, final_url = self.post_form("/login", [("_csrf", csrf), ("username", username), ("password", password)])
        if "/login" in final_url:
            raise RuntimeError(f"网页登录失败：{final_url}")

    def api_request(self, method: str, path: str, token: Optional[str] = None, body: Optional[dict] = None) -> dict:
        headers = {}
        data = None
        if token:
            headers["Authorization"] = f"Bearer {token}"
        if body is not None:
            headers["Content-Type"] = "application/json;charset=UTF-8"
            data = json.dumps(body, ensure_ascii=False).encode("utf-8")
        request = urllib.request.Request(self._url(path), data=data, headers=headers, method=method.upper())
        with self._open(urllib.request.urlopen, request, 180) as response:
            return json.loads(response.read().decode("utf-8", errors="replace"))

    def api_login(self, username: str, password: str) -> dict:
        payload = self.api_request("POST", "/api/auth/login", body={"username": username, "password": password})
        if payload.get("code") != 200 or not payload.get("data", {}).get("token"):
            raise RuntimeError(f"API 登录失败：{payload}")
        return payload


class NoRedirectHandler(urllib.request.HTTPRedirectHandler):
    def http_error_301(self, req, fp, code, msg, headers):
        return fp

    def http_error_302(self, req, fp, code, msg, headers):
        return fp

    def http_error_303(self, req, fp, code, msg, headers):
        return fp

    def http_error_307(self, req, fp, code, msg, headers):
        return fp

    def http_error_308(self, req, fp, code, msg, headers):
        return fp


def load_accounts(path: Path) -> List[Account]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = []
        for row in reader:
            rows.append(
                Account(
                    username=row["username"].strip(),
                    password=row["password"].strip(),
                    student_no=row["student_no"].strip(),
                    student_name=row["student_name"].strip(),
                )
            )
        return rows


def infer_prefix(accounts: List[Account]) -> str:
    sample = accounts[0].student_name
    parts = sample.split("-")
    if len(parts) < 3:
        raise RuntimeError("无法从学生姓名推断测试前缀")
    return parts[0]


def candidate_passwords(current_password: str) -> List[str]:
    second = next_password(current_password)
    third = next_password(second)
    return [current_password, second, third]


def api_login_student(account: Account) -> Tuple[RemoteClient, dict]:
    last_error = None
    for password in candidate_passwords(account.password):
        client = RemoteClient(BASE_URL)
        try:
            payload = client.api_login(account.username, password)
            account.password = password
            return client, payload
        except Exception as exc:
            last_error = exc
    raise last_error


def web_login_student(account: Account) -> RemoteClient:
    last_error = None
    for password in candidate_passwords(account.password):
        client = RemoteClient(BASE_URL)
        try:
            client.login_web(account.username, password)
            account.password = password
            return client
        except Exception as exc:
            last_error = exc
    raise last_error


def fetch_classes(client: RemoteClient, admin_token: str) -> List[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=admin_token)
    result = []
    for item in payload.get("data", []):
        if item.get("gradeName") == GRADE_NAME and item.get("name") in CLASS_NAMES and item.get("type") == "行政班":
            result.append(item)
    result.sort(key=lambda item: int(str(item["name"]).replace("班", "")))
    if len(result) != len(CLASS_NAMES):
        raise RuntimeError(f"仅找到 {len(result)} 个目标行政班")
    return result


def fetch_target_students(client: RemoteClient, admin_token: str, classes: List[dict], prefix: str, expected_count: int) -> List[dict]:
    results = []
    for class_row in classes:
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=100",
            token=admin_token,
        )
        content = payload.get("data", {}).get("content", [])
        matches = [item for item in content if str(item.get("name", "")).startswith(prefix)]
        results.extend(matches)
    if len(results) != expected_count:
        raise RuntimeError(f"目标学生数量为 {len(results)}，预期 {expected_count}")
    return results


def reset_student_passwords(client: RemoteClient, student_ids: List[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick("/admin/student-accounts/reset-password", fields)


def refresh_accounts_from_export(client: RemoteClient, output_path: Path) -> List[Account]:
    output_path.write_bytes(client.get_bytes("/admin/student-accounts/export"))
    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_map = {str(value): index for index, value in enumerate(rows[0])}
    accounts = []
    for row in rows[1:]:
        if not row or not row[1] or not row[4] or not row[5]:
            continue
        accounts.append(
            Account(
                username=str(row[4]).strip(),
                password=str(row[5]).strip(),
                student_no=str(row[1]).strip(),
                student_name=str(row[0]).strip(),
            )
        )
    return accounts


def create_event(client: RemoteClient, event_name: str, csrf: str) -> int:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(minutes=10)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now + timedelta(days=1)).strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form_quick("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"创建活动失败，跳转地址：{final_url}")
    return int(match.group(1))


def save_event_students(client: RemoteClient, event_id: int, student_ids: List[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick(f"/admin/courses/{event_id}/students/save", fields)


def create_courses(client: RemoteClient, event_id: int, class_rows: List[dict], csrf: str) -> None:
    class_ids = [str(item["id"]) for item in class_rows]
    for name, mode, capacity in COURSE_SPECS:
        fields = [
            ("_csrf", csrf),
            ("name", name),
            ("description", f"{name} 自动测试课程"),
            ("capacityMode", mode),
        ]
        if mode == "GLOBAL":
            fields.append(("totalCapacity", str(capacity)))
        else:
            for class_id in class_ids:
                fields.append(("classIds", class_id))
                fields.append(("classCapacities", str(capacity)))
        client.post_form_quick(f"/admin/courses/{event_id}/courses/save", fields)


def activate_all_courses(client: RemoteClient, event_id: int, csrf: str) -> str:
    page, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    course_ids = [int(match) for match in re.findall(rf"/admin/courses/{event_id}/courses/(\d+)/activate", page)]
    for course_id in course_ids:
        client.post_form_quick(f"/admin/courses/{event_id}/courses/{course_id}/activate", [("_csrf", csrf)])
    return csrf


def start_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])


def fetch_course_map_for_current_event(account: Account) -> Dict[str, int]:
    client, login = api_login_student(account)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/courses", token=token)
    course_map = {}
    for item in payload.get("data", []):
        course_map[str(item["name"])] = int(item["id"])
    return course_map


def submit_preferences(account: Account, course_map: Dict[str, int]) -> Dict[str, str]:
    client, login = api_login_student(account)
    token = login["data"]["token"]
    pref1_id = course_map[ROUND1_PREF1]
    pref2_id = course_map[ROUND1_PREF2]
    result1 = client.api_request("POST", f"/api/student/courses/{pref1_id}/prefer?preference=1", token=token)
    result2 = client.api_request("POST", f"/api/student/courses/{pref2_id}/prefer?preference=2", token=token)
    return {
        "username": account.username,
        "pref1Code": str(result1.get("code")),
        "pref1Message": str(result1.get("message")),
        "pref2Code": str(result2.get("code")),
        "pref2Message": str(result2.get("message")),
    }


def activate_and_confirm_round1(account: Account) -> Dict[str, str]:
    client = web_login_student(account)

    password_page, _ = client.get_text("/student/password?force=true")
    password_csrf = parse_csrf_token(password_page)
    new_password = next_password(account.password)
    client.post_form_quick(
        "/student/password",
        [
            ("_csrf", password_csrf),
            ("oldPassword", account.password),
            ("newPassword", new_password),
            ("confirmPassword", new_password),
            ("force", "true"),
        ],
    )
    account.password = new_password

    courses_page, _ = client.get_text("/student/courses")
    courses_csrf = parse_csrf_token(courses_page)
    client.post_form_quick("/student/courses/confirm", [("_csrf", courses_csrf)])
    return {
        "username": account.username,
        "student_no": account.student_no,
        "student_name": account.student_name,
        "new_password": account.password,
        "status": "OK",
    }


def process_round1(client: RemoteClient, event_id: int, csrf: str) -> float:
    start = time.perf_counter()
    client.post_form_quick(f"/admin/courses/events/{event_id}/process", [("_csrf", csrf)])
    for _ in range(180):
        payload, _ = client.get_text(f"/admin/courses/events/{event_id}/lottery-status")
        data = json.loads(payload)
        status = data.get("status")
        if status and status != "PROCESSING":
            return (time.perf_counter() - start) * 1000
        time.sleep(1)
    raise RuntimeError("等待第一轮结算超时")


def fetch_student_selection_result(account: Account) -> Dict[str, object]:
    client, login = api_login_student(account)
    token = login["data"]["token"]
    payload = client.api_request("GET", "/api/student/my-selections", token=token)
    selections = payload.get("data", [])
    confirmed = next((item for item in selections if item.get("status") == "CONFIRMED"), None)
    return {
        "account": account,
        "selections": selections,
        "confirmed": confirmed,
    }


def second_round_select(account: Account, course_id: int, barrier: threading.Barrier) -> Dict[str, object]:
    client = RemoteClient(BASE_URL)
    token = client.api_login(account.username, account.password)["data"]["token"]
    barrier.wait(timeout=20)
    started_at = time.perf_counter()
    try:
        payload = client.api_request("POST", f"/api/student/courses/{course_id}/select", token=token)
        elapsed_ms = (time.perf_counter() - started_at) * 1000
        return {
            "username": account.username,
            "student_no": account.student_no,
            "student_name": account.student_name,
            "code": int(payload.get("code", 0)),
            "message": str(payload.get("message", "")),
            "elapsedMs": round(elapsed_ms, 2),
        }
    except Exception as exc:
        elapsed_ms = (time.perf_counter() - started_at) * 1000
        return {
            "username": account.username,
            "student_no": account.student_no,
            "student_name": account.student_name,
            "code": -1,
            "message": str(exc),
            "elapsedMs": round(elapsed_ms, 2),
        }


def fetch_round2_token(account: Account) -> Dict[str, object]:
    client, login = api_login_student(account)
    token = login["data"]["token"]
    return {
        "account": account,
        "token": token,
    }


def second_round_select_with_token(account: Account, token: str, course_id: int, start_event: threading.Event) -> Dict[str, object]:
    client = RemoteClient(BASE_URL)
    start_event.wait(timeout=30)
    started_at = time.perf_counter()
    try:
        payload = client.api_request("POST", f"/api/student/courses/{course_id}/select", token=token)
        elapsed_ms = (time.perf_counter() - started_at) * 1000
        return {
            "username": account.username,
            "student_no": account.student_no,
            "student_name": account.student_name,
            "code": int(payload.get("code", 0)),
            "message": str(payload.get("message", "")),
            "elapsedMs": round(elapsed_ms, 2),
        }
    except Exception as exc:
        elapsed_ms = (time.perf_counter() - started_at) * 1000
        return {
            "username": account.username,
            "student_no": account.student_no,
            "student_name": account.student_name,
            "code": -1,
            "message": str(exc),
            "elapsedMs": round(elapsed_ms, 2),
        }


def write_csv(path: Path, rows: List[Dict[str, object]], headers: List[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def main() -> None:
    started = datetime.now()
    ensure_dir(RESULT_DIR)
    accounts = load_accounts(ACCOUNTS_CSV)
    prefix = infer_prefix(accounts)
    run_tag = started.strftime("%Y%m%d_%H%M%S")
    run_dir = RESULT_DIR / f"round1_round2_{run_tag}"
    ensure_dir(run_dir)

    admin_client = RemoteClient(BASE_URL)
    admin_client.login_web(ADMIN_USERNAME, ADMIN_PASSWORD)
    admin_page, _ = admin_client.get_text("/admin/courses")
    admin_csrf = parse_csrf_token(admin_page)
    admin_token = admin_client.api_login(ADMIN_USERNAME, ADMIN_PASSWORD)["data"]["token"]

    classes = fetch_classes(admin_client, admin_token)
    students = fetch_target_students(admin_client, admin_token, classes, prefix, len(accounts))
    student_ids = [int(item["id"]) for item in students]
    reset_student_passwords(admin_client, student_ids, admin_csrf)
    exported_accounts = refresh_accounts_from_export(admin_client, run_dir / "accounts_reset.xlsx")
    account_map = {account.student_no: account for account in exported_accounts}
    accounts = [account_map[account.student_no] for account in accounts if account.student_no in account_map]
    if len(accounts) != len(student_ids):
        raise RuntimeError(f"重置并导出后的账号数为 {len(accounts)}，预期 {len(student_ids)}")
    event_name = f"{started.strftime('%Y-%m-%d %H:%M:%S')} 第一轮抽签+第二轮并发测试 {prefix}"
    event_id = create_event(admin_client, event_name, admin_csrf)
    save_event_students(admin_client, event_id, student_ids, admin_csrf)
    create_courses(admin_client, event_id, classes, admin_csrf)
    admin_csrf = activate_all_courses(admin_client, event_id, admin_csrf)
    start_round1(admin_client, event_id, admin_csrf)

    sample_course_map = fetch_course_map_for_current_event(accounts[0])
    pref_submit_rows: List[Dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=PREFERENCE_WORKERS) as executor:
        futures = [executor.submit(submit_preferences, account, sample_course_map) for account in accounts]
        for future in as_completed(futures):
            pref_submit_rows.append(future.result())
    write_csv(
        run_dir / "round1_preference_submit.csv",
        pref_submit_rows,
        ["username", "pref1Code", "pref1Message", "pref2Code", "pref2Message"],
    )

    confirm_rows: List[Dict[str, str]] = []
    with ThreadPoolExecutor(max_workers=CONFIRM_WORKERS) as executor:
        futures = [executor.submit(activate_and_confirm_round1, account) for account in accounts]
        for future in as_completed(futures):
            confirm_rows.append(future.result())
    write_csv(
        run_dir / "round1_confirm_submit.csv",
        confirm_rows,
        ["username", "student_no", "student_name", "new_password", "status"],
    )

    lottery_duration_ms = process_round1(admin_client, event_id, admin_csrf)

    selection_results: List[Dict[str, object]] = []
    with ThreadPoolExecutor(max_workers=RESULT_WORKERS) as executor:
        futures = [executor.submit(fetch_student_selection_result, account) for account in accounts]
        for future in as_completed(futures):
            selection_results.append(future.result())

    first_choice_confirmed = 0
    second_choice_confirmed = 0
    unsuccessful_accounts: List[Account] = []
    round1_course_counts: Dict[str, int] = {}
    per_student_rows: List[Dict[str, object]] = []

    for item in selection_results:
        account = item["account"]
        confirmed = item["confirmed"]
        selections = item["selections"]
        if confirmed:
            preference = int(confirmed.get("preference", 0))
            course_name = str(confirmed.get("courseName"))
            round1_course_counts[course_name] = round1_course_counts.get(course_name, 0) + 1
            if preference == 1:
                first_choice_confirmed += 1
            elif preference == 2:
                second_choice_confirmed += 1
            status = "CONFIRMED"
        else:
            unsuccessful_accounts.append(account)
            status = "UNSUCCESSFUL"
        per_student_rows.append(
            {
                "username": account.username,
                "student_no": account.student_no,
                "student_name": account.student_name,
                "status": status,
                "confirmed_course": confirmed.get("courseName") if confirmed else "",
                "confirmed_preference": confirmed.get("preference") if confirmed else "",
                "selection_json": json.dumps(selections, ensure_ascii=False),
            }
        )

    unsuccessful_count = len(unsuccessful_accounts)
    write_csv(
        run_dir / "round1_student_results.csv",
        per_student_rows,
        ["username", "student_no", "student_name", "status", "confirmed_course", "confirmed_preference", "selection_json"],
    )

    round2_users = min(ROUND2_MAX_USERS, unsuccessful_count)
    round2_accounts = unsuccessful_accounts[:round2_users]
    round2_course_id = sample_course_map[ROUND2_TARGET]
    round2_tokens: List[Dict[str, object]] = []
    with ThreadPoolExecutor(max_workers=min(round2_users, RESULT_WORKERS)) as executor:
        futures = [executor.submit(fetch_round2_token, account) for account in round2_accounts]
        for future in as_completed(futures):
            round2_tokens.append(future.result())
    start_event = threading.Event()
    round2_started_at = time.perf_counter()
    round2_rows: List[Dict[str, object]] = []
    with ThreadPoolExecutor(max_workers=round2_users) as executor:
        futures = [
            executor.submit(second_round_select_with_token, item["account"], item["token"], round2_course_id, start_event)
            for item in round2_tokens
        ]
        start_event.set()
        for future in as_completed(futures):
            round2_rows.append(future.result())
    round2_wall_ms = (time.perf_counter() - round2_started_at) * 1000
    write_csv(
        run_dir / "round2_concurrency_results.csv",
        round2_rows,
        ["username", "student_no", "student_name", "code", "message", "elapsedMs"],
    )

    round2_success = [row for row in round2_rows if int(row["code"]) == 200]
    round2_failure = [row for row in round2_rows if int(row["code"]) != 200]
    latencies = [float(row["elapsedMs"]) for row in round2_rows]
    round2_target_capacity = COURSE_CAPACITY[ROUND2_TARGET]
    round2_success_users = sorted({str(row["username"]) for row in round2_success})
    round2_failure_messages = sorted({str(row["message"]) for row in round2_failure})

    final_selection_results: List[Dict[str, object]] = []
    with ThreadPoolExecutor(max_workers=RESULT_WORKERS) as executor:
        futures = [executor.submit(fetch_student_selection_result, account) for account in accounts]
        for future in as_completed(futures):
            final_selection_results.append(future.result())

    final_course_counts: Dict[str, int] = {}
    final_target_confirmed = 0
    total_confirmed_after_round2 = 0
    for item in final_selection_results:
        confirmed = item["confirmed"]
        if not confirmed:
            continue
        total_confirmed_after_round2 += 1
        course_name = str(confirmed.get("courseName"))
        final_course_counts[course_name] = final_course_counts.get(course_name, 0) + 1
        if course_name == ROUND2_TARGET:
            final_target_confirmed += 1

    summary = {
        "baseUrl": BASE_URL,
        "eventId": event_id,
        "eventName": event_name,
        "studentPrefix": prefix,
        "studentCount": len(accounts),
        "firstRound": {
            "preference1": ROUND1_PREF1,
            "preference2": ROUND1_PREF2,
            "lotteryDurationMs": round(lottery_duration_ms, 2),
            "firstChoiceConfirmed": first_choice_confirmed,
            "secondChoiceConfirmed": second_choice_confirmed,
            "unsuccessful": unsuccessful_count,
            "courseConfirmedCounts": round1_course_counts,
        },
        "secondRound": {
            "targetCourse": ROUND2_TARGET,
            "targetCourseId": round2_course_id,
            "targetCourseCapacity": round2_target_capacity,
            "concurrentUsers": round2_users,
            "successCount": len(round2_success),
            "successUserCount": len(round2_success_users),
            "failureCount": len(round2_failure),
            "wallTimeMs": round(round2_wall_ms, 2),
            "latencyMs": {
                "min": round(min(latencies), 2) if latencies else 0,
                "avg": round(statistics.mean(latencies), 2) if latencies else 0,
                "p50": round(percentile(latencies, 0.50), 2) if latencies else 0,
                "p90": round(percentile(latencies, 0.90), 2) if latencies else 0,
                "p95": round(percentile(latencies, 0.95), 2) if latencies else 0,
                "max": round(max(latencies), 2) if latencies else 0,
            },
            "sampleFailures": round2_failure_messages[:5],
            "finalConfirmedInTargetCourse": final_target_confirmed,
            "oversold": final_target_confirmed > round2_target_capacity,
            "oversoldBy": max(0, final_target_confirmed - round2_target_capacity),
        },
        "finalConfirmedCounts": final_course_counts,
        "totalConfirmedAfterRound2": total_confirmed_after_round2,
        "files": {
            "accountsResetXlsx": str(run_dir / "accounts_reset.xlsx"),
            "round1PreferenceSubmit": str(run_dir / "round1_preference_submit.csv"),
            "round1ConfirmSubmit": str(run_dir / "round1_confirm_submit.csv"),
            "round1StudentResults": str(run_dir / "round1_student_results.csv"),
            "round2ConcurrencyResults": str(run_dir / "round2_concurrency_results.csv"),
            "summaryJson": str(run_dir / "summary.json"),
        },
    }

    summary_path = run_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
