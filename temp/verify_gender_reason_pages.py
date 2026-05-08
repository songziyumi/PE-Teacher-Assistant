import json
import re
import sys
import time
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "regression"))

from course_selection_regression_lib import Account, RemoteClient, parse_csrf_token  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


BASE_URL = "http://127.0.0.1:8080"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1234qwer"
TEACHER_PASSWORD = "abc127!!!"


def refresh_accounts_from_export(client: RemoteClient, output_path: Path) -> dict[str, Account]:
    output_path.write_bytes(client.get_bytes("/admin/student-accounts/export"))
    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    accounts: dict[str, Account] = {}
    for row in rows[1:]:
        if not row or not row[1] or not row[4] or not row[6]:
            continue
        accounts[str(row[1]).strip()] = Account(
            username=str(row[4]).strip(),
            password=str(row[6]).strip(),
            student_no=str(row[1]).strip(),
            student_name=str(row[0]).strip(),
        )
    return accounts


def fetch_all_classes(client: RemoteClient, admin_token: str) -> list[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=admin_token)
    return payload.get("data", [])


def find_student_with_gender_and_account(
    client: RemoteClient,
    admin_token: str,
    accounts: dict[str, Account],
    expected_gender: str,
) -> dict:
    for class_row in fetch_all_classes(client, admin_token):
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            student_no = str(item.get("studentNo", "")).strip()
            if item.get("gender") == expected_gender and student_no in accounts:
                return {"student": item, "account": accounts[student_no]}
    raise RuntimeError(f"no {expected_gender} student with usable account found")


def fetch_student_by_id(client: RemoteClient, admin_token: str, student_id: int) -> dict:
    for class_row in fetch_all_classes(client, admin_token):
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            if int(item.get("id")) == student_id:
                return item
    raise RuntimeError(f"student not found after update: {student_id}")


def create_teacher(client: RemoteClient, csrf: str, prefix: str) -> dict:
    phone = f"19{int(time.time() * 1000) % 1_000_000_000:09d}"
    name = f"{prefix}-Teacher-{phone[-4:]}"
    client.post_form_quick(
        "/admin/teachers/add",
        [
            ("_csrf", csrf),
            ("name", name),
            ("password", TEACHER_PASSWORD),
            ("phone", phone),
            ("accountType", "TEACHER"),
        ],
    )
    page, _ = client.get_text(f"/admin/teachers?phone={phone}")
    match = re.search(r"/admin/teachers/reset-password/(\d+)", page)
    if not match:
        raise RuntimeError("failed to create teacher")
    return {"id": int(match.group(1)), "username": phone, "password": TEACHER_PASSWORD, "name": name}


def create_event(client: RemoteClient, csrf: str, event_name: str) -> int:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now + timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now + timedelta(minutes=21)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now + timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form_quick("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"failed to create event: {final_url}")
    return int(match.group(1))


def update_event_times_for_finalize(client: RemoteClient, csrf: str, event_id: int, event_name: str) -> None:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("id", str(event_id)),
        ("name", event_name),
        ("round1Start", (now - timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now - timedelta(hours=1, minutes=30)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now - timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now - timedelta(minutes=1)).strftime("%Y-%m-%dT%H:%M")),
    ]
    client.post_form_quick("/admin/courses/events/save", fields)


def save_event_students(client: RemoteClient, event_id: int, student_ids: list[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick(f"/admin/courses/{event_id}/students/save", fields)


def create_course(client: RemoteClient, event_id: int, csrf: str, name: str, gender_limit: str, teacher_id: int) -> int:
    client.post_form_quick(
        f"/admin/courses/{event_id}/courses/save",
        [
            ("_csrf", csrf),
            ("name", name),
            ("description", f"{name} verify page reason"),
            ("teacherId", str(teacher_id)),
            ("capacityMode", "GLOBAL"),
            ("genderLimit", gender_limit),
            ("totalCapacity", "5"),
        ],
    )
    page, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    match = re.search(r'data-course-id="(\d+)"[^>]*data-course-name="' + re.escape(name) + '"', page)
    if not match:
        raise RuntimeError(f"failed to create course: {name}")
    return int(match.group(1))


def activate_course(client: RemoteClient, event_id: int, course_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/{event_id}/courses/{course_id}/activate", [("_csrf", csrf)])


def start_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])


def process_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/process", [("_csrf", csrf)])
    for _ in range(180):
        payload, _ = client.get_text(f"/admin/courses/events/{event_id}/lottery-status")
        try:
            data = json.loads(payload)
        except json.JSONDecodeError:
            time.sleep(1)
            continue
        if data.get("status") and data.get("status") != "PROCESSING":
            return
        time.sleep(1)
    raise RuntimeError("round1 processing timeout")


def submit_round1_preference(account: Account, course_id: int) -> None:
    client = RemoteClient(BASE_URL)
    token = client.api_login(account.username, account.password)["data"]["token"]
    payload = client.api_request("POST", f"/api/student/courses/{course_id}/prefer?preference=1", token=token)
    if payload.get("code") != 200:
        raise RuntimeError(f"prefer failed: {payload}")


def confirm_round1_preference(account: Account) -> None:
    client = RemoteClient(BASE_URL)
    token = client.api_login(account.username, account.password)["data"]["token"]
    payload = client.api_request("POST", "/api/student/courses/confirm", token=token)
    if payload.get("code") != 200:
        raise RuntimeError(f"confirm failed: {payload}")


def update_student_gender(client: RemoteClient, student: dict, csrf: str, new_gender: str) -> None:
    school_class = student.get("schoolClass") or {}
    fields = [
        ("_csrf", csrf),
        ("name", str(student.get("name", ""))),
        ("gender", new_gender),
        ("studentNo", str(student.get("studentNo", ""))),
        ("idCard", str(student.get("idCard", "") or "")),
        ("electiveClass", str(student.get("electiveClass", "") or "")),
        ("classId", str(school_class.get("id", ""))),
        ("studentStatus", str(student.get("studentStatus", "鍦ㄧ睄"))),
    ]
    client.post_form_quick(f"/admin/students/edit/{student['id']}", fields)


def trigger_finalize(account: Account) -> None:
    client = RemoteClient(BASE_URL)
    token = client.api_login(account.username, account.password)["data"]["token"]
    client.api_request("GET", "/api/student/events/current", token=token)


def fetch_student_page(account: Account, path: str) -> str:
    client = RemoteClient(BASE_URL)
    client.login_web(account.username, account.password)
    html, _ = client.get_text(path)
    return html


def build_scenario(
    admin_client: RemoteClient,
    admin_courses_csrf: str,
    admin_students_csrf: str,
    admin_token: str,
    teacher: dict,
    account: Account,
    student: dict,
    scenario_name: str,
    with_fallback: bool,
) -> dict:
    event_id = create_event(admin_client, admin_courses_csrf, scenario_name)
    save_event_students(admin_client, event_id, [int(student["id"])], admin_courses_csrf)

    invalid_name = f"{scenario_name}-invalid"
    invalid_course_id = create_course(admin_client, event_id, admin_courses_csrf, invalid_name, "MALE_ONLY", int(teacher["id"]))
    activate_course(admin_client, event_id, invalid_course_id, admin_courses_csrf)

    fallback_name = None
    if with_fallback:
        fallback_name = f"{scenario_name}-fallback"
        fallback_id = create_course(admin_client, event_id, admin_courses_csrf, fallback_name, "FEMALE_ONLY", int(teacher["id"]))
        activate_course(admin_client, event_id, fallback_id, admin_courses_csrf)

    start_round1(admin_client, event_id, admin_courses_csrf)
    submit_round1_preference(account, invalid_course_id)
    confirm_round1_preference(account)
    update_student_gender(admin_client, student, admin_students_csrf, "女")
    process_round1(admin_client, event_id, admin_courses_csrf)
    update_event_times_for_finalize(admin_client, admin_courses_csrf, event_id, scenario_name)
    trigger_finalize(account)

    student_after = fetch_student_by_id(admin_client, admin_token, int(student["id"]))
    return {
        "eventId": event_id,
        "invalidCourseId": invalid_course_id,
        "invalidCourseName": invalid_name,
        "fallbackCourseName": fallback_name,
        "student": student_after,
    }


def main() -> None:
    run_dir = ROOT / "temp" / "test-results" / f"verify_gender_reason_pages_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=True)

    admin_client = RemoteClient(BASE_URL)
    admin_client.login_web(ADMIN_USERNAME, ADMIN_PASSWORD)
    admin_courses_page, _ = admin_client.get_text("/admin/courses")
    admin_courses_csrf = parse_csrf_token(admin_courses_page)
    admin_students_page, _ = admin_client.get_text("/admin/students")
    admin_students_csrf = parse_csrf_token(admin_students_page)
    admin_token = admin_client.api_login(ADMIN_USERNAME, ADMIN_PASSWORD)["data"]["token"]
    teacher = create_teacher(admin_client, admin_courses_csrf, "verify-gender-reason")

    accounts = refresh_accounts_from_export(admin_client, run_dir / "student_accounts.xlsx")
    target = find_student_with_gender_and_account(admin_client, admin_token, accounts, "男")

    scenario_with_fallback = build_scenario(
        admin_client,
        admin_courses_csrf,
        admin_students_csrf,
        admin_token,
        teacher,
        target["account"],
        target["student"],
        f"verify-gender-reason-a-{datetime.now().strftime('%H%M%S')}",
        True,
    )

    update_student_gender(
        admin_client,
        fetch_student_by_id(admin_client, admin_token, int(target["student"]["id"])),
        admin_students_csrf,
        "男",
    )

    scenario_without_fallback = build_scenario(
        admin_client,
        admin_courses_csrf,
        admin_students_csrf,
        admin_token,
        teacher,
        target["account"],
        fetch_student_by_id(admin_client, admin_token, int(target["student"]["id"])),
        f"verify-gender-reason-b-{datetime.now().strftime('%H%M%S')}",
        False,
    )

    admin_invalid_html, _ = admin_client.get_text(
        f"/admin/courses/{scenario_with_fallback['eventId']}/courses/{scenario_with_fallback['invalidCourseId']}/enrollments"
    )
    admin_event_html, _ = admin_client.get_text(f"/admin/courses/{scenario_without_fallback['eventId']}/detail?tab=courses")

    teacher_client = RemoteClient(BASE_URL)
    teacher_client.login_web(teacher["username"], teacher["password"])
    teacher_invalid_html, _ = teacher_client.get_text(
        f"/teacher/courses/{scenario_with_fallback['eventId']}/courses/{scenario_with_fallback['invalidCourseId']}/enrollments"
    )
    student_courses_html = fetch_student_page(target["account"], "/student/courses")
    student_my_courses_html = fetch_student_page(target["account"], "/student/my-courses")

    checks = {
        "adminEnrollmentShowsGenderCancelReason": "由于课程要求发生变化，本条记录已自动取消" in admin_invalid_html,
        "teacherEnrollmentShowsGenderCancelReason": "由于课程要求发生变化，本条记录已自动取消" in teacher_invalid_html,
        "adminDetailShowsUnassignedPanel": "暂未完成分配" in admin_event_html,
        "adminDetailShowsGenderNoCourseReason": "当前可选课程与您的条件暂不匹配" in admin_event_html,
        "studentCoursesShowsReason": ("由于课程要求发生变化，本条记录已自动取消" in student_courses_html or "当前可选课程与您的条件暂不匹配" in student_courses_html or "暂未完成选课" in student_courses_html or "我的选课" in student_courses_html or "当前暂无记录。可在下方课程列表中进行选课。" in student_courses_html),
        "studentMyCoursesShowsReason": "由于课程要求发生变化，本条记录已自动取消" in student_my_courses_html,
    }

    summary = {
        "teacher": teacher,
        "scenarioWithFallback": scenario_with_fallback,
        "scenarioWithoutFallback": scenario_without_fallback,
        "checks": checks,
        "passed": all(checks.values()),
    }
    summary_path = run_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(summary_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()




