import json
import re
import sys
from datetime import datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts" / "regression"))

from course_selection_regression_lib import Account, RemoteClient, parse_csrf_token  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


BASE_URL = "http://127.0.0.1:8080"
ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "1234qwer"


def refresh_accounts_from_export(client: RemoteClient, output_path: Path) -> dict[str, Account]:
    output_path.write_bytes(client.get_bytes("/admin/student-accounts/export"))
    workbook = load_workbook(output_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    accounts: dict[str, Account] = {}
    for row in rows[1:]:
        if not row or not row[1] or not row[4] or not row[6]:
            continue
        account = Account(
            username=str(row[4]).strip(),
            password=str(row[6]).strip(),
            student_no=str(row[1]).strip(),
            student_name=str(row[0]).strip(),
        )
        accounts[account.student_no] = account
    return accounts


def fetch_all_classes(client: RemoteClient, admin_token: str) -> list[dict]:
    payload = client.api_request("GET", "/api/admin/classes", token=admin_token)
    return payload.get("data", [])


def find_male_student_with_account(client: RemoteClient, admin_token: str, accounts: dict[str, Account]) -> dict:
    for class_row in fetch_all_classes(client, admin_token):
        payload = client.api_request(
            "GET",
            f"/api/admin/students?classId={class_row['id']}&page=0&size=200",
            token=admin_token,
        )
        for item in payload.get("data", {}).get("content", []):
            student_no = str(item.get("studentNo", "")).strip()
            if item.get("gender") == "男" and student_no in accounts:
                return {"student": item, "account": accounts[student_no]}
    raise RuntimeError("no male student with usable account found")


def create_event(client: RemoteClient, csrf: str, event_name: str) -> int:
    now = datetime.now()
    fields = [
        ("_csrf", csrf),
        ("name", event_name),
        ("round1Start", (now - timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round1End", (now + timedelta(minutes=20)).strftime("%Y-%m-%dT%H:%M")),
        ("round2Start", (now + timedelta(minutes=21)).strftime("%Y-%m-%dT%H:%M")),
        ("round2End", (now + timedelta(hours=1)).strftime("%Y-%m-%dT%H:%M")),
    ]
    _, final_url = client.post_form_quick("/admin/courses/events/save", fields)
    match = re.search(r"/admin/courses/(\d+)/detail", final_url)
    if not match:
        raise RuntimeError(f"failed to create event: {final_url}")
    return int(match.group(1))


def save_event_students(client: RemoteClient, event_id: int, student_ids: list[int], csrf: str) -> None:
    fields = [("_csrf", csrf)]
    fields.extend(("studentIds", str(student_id)) for student_id in student_ids)
    client.post_form_quick(f"/admin/courses/{event_id}/students/save", fields)


def create_course(client: RemoteClient, event_id: int, csrf: str, name: str) -> int:
    client.post_form_quick(
        f"/admin/courses/{event_id}/courses/save",
        [
            ("_csrf", csrf),
            ("name", name),
            ("description", f"{name} guard test"),
            ("capacityMode", "GLOBAL"),
            ("genderLimit", "ALL"),
            ("totalCapacity", "5"),
        ],
    )
    page, _ = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    match = re.search(r'data-course-id="(\d+)"[^>]*data-course-name="' + re.escape(name) + '"', page)
    if not match:
        raise RuntimeError("failed to find created course id")
    return int(match.group(1))


def activate_course(client: RemoteClient, event_id: int, course_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/{event_id}/courses/{course_id}/activate", [("_csrf", csrf)])


def start_round1(client: RemoteClient, event_id: int, csrf: str) -> None:
    client.post_form_quick(f"/admin/courses/events/{event_id}/start-round1", [("_csrf", csrf)])


def submit_round1_preference(account: Account, course_id: int) -> dict:
    client = RemoteClient(BASE_URL)
    login = client.api_login(account.username, account.password)
    token = login["data"]["token"]
    return client.api_request("POST", f"/api/student/courses/{course_id}/prefer?preference=1", token=token)


def attempt_gender_limit_update(client: RemoteClient, event_id: int, course_id: int, course_name: str, csrf: str) -> tuple[str, str]:
    client.post_form_quick(
        f"/admin/courses/{event_id}/courses/save",
        [
            ("_csrf", csrf),
            ("courseId", str(course_id)),
            ("name", course_name),
            ("description", f"{course_name} guard test"),
            ("capacityMode", "GLOBAL"),
            ("genderLimit", "FEMALE_ONLY"),
            ("totalCapacity", "5"),
        ],
    )
    page, final_url = client.get_text(f"/admin/courses/{event_id}/detail?tab=courses")
    return page, final_url


def parse_course_gender_limit(detail_html: str, course_id: int) -> str | None:
    match = re.search(
        rf'data-course-id="{course_id}"[^>]*data-course-gender-limit="([^"]+)"',
        detail_html,
    )
    return match.group(1) if match else None


def main() -> None:
    run_dir = ROOT / "temp" / "test-results" / f"gender_limit_edit_guard_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    run_dir.mkdir(parents=True, exist_ok=True)

    admin_client = RemoteClient(BASE_URL)
    admin_client.login_web(ADMIN_USERNAME, ADMIN_PASSWORD)
    admin_page, _ = admin_client.get_text("/admin/courses")
    admin_csrf = parse_csrf_token(admin_page)
    admin_token = admin_client.api_login(ADMIN_USERNAME, ADMIN_PASSWORD)["data"]["token"]

    accounts = refresh_accounts_from_export(admin_client, run_dir / "student_accounts.xlsx")
    male_target = find_male_student_with_account(admin_client, admin_token, accounts)

    event_name = f"Gender-Edit-Guard-{datetime.now().strftime('%H%M%S')}"
    course_name = f"Guard-Course-{datetime.now().strftime('%H%M%S')}"
    event_id = create_event(admin_client, admin_csrf, event_name)
    save_event_students(admin_client, event_id, [int(male_target["student"]["id"])], admin_csrf)
    course_id = create_course(admin_client, event_id, admin_csrf, course_name)
    activate_course(admin_client, event_id, course_id, admin_csrf)
    start_round1(admin_client, event_id, admin_csrf)

    prefer_result = submit_round1_preference(male_target["account"], course_id)
    detail_html, detail_url = attempt_gender_limit_update(admin_client, event_id, course_id, course_name, admin_csrf)
    actual_gender_limit = parse_course_gender_limit(detail_html, course_id)

    passed = (
        prefer_result.get("code") == 200
        and detail_url.endswith("?tab=courses")
        and actual_gender_limit == "ALL"
    )

    summary = {
        "eventId": event_id,
        "courseId": course_id,
        "studentId": male_target["student"]["id"],
        "studentNo": male_target["account"].student_no,
        "studentName": male_target["account"].student_name,
        "preferResult": prefer_result,
        "detailUrl": detail_url,
        "actualGenderLimitAfterSaveAttempt": actual_gender_limit,
        "saveBlocked": actual_gender_limit == "ALL",
        "passed": passed,
    }
    summary_path = run_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(summary_path)
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
